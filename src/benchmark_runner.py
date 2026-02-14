# src/benchmark_runner.py
"""
Benchmark runner for multiple instances with Excel output.
"""
import time
from pathlib import Path
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .instance import load_ortec_vrptw
from .solver import solve_instance


def run_benchmark(
    instance_paths: List[Path],
    time_budgets: List[int],
    output_excel: Path,
    validate: bool = True,
    verbose: bool = False,
) -> None:
    """
    Run benchmark on multiple instances with multiple time budgets.
    
    Args:
        instance_paths: List of instance file paths
        time_budgets: List of time budgets in seconds (e.g., [300, 600])
        output_excel: Path for output Excel file
        validate: Perform validation
        verbose: Print detailed progress
    """
    results = []
    
    for inst_path in instance_paths:
        inst_name = inst_path.stem
        print(f"\n{'='*80}")
        print(f"Instance: {inst_name}")
        print(f"{'='*80}")
        
        # Load instance
        print(f"Loading instance...")
        inst = load_ortec_vrptw(inst_path)
        print(f"  Nodes: {inst.n_nodes}, Vehicles: {inst.n_vehicles}, Capacity: {inst.capacity}")
        
        for budget in time_budgets:
            print(f"\n--- Budget: {budget}s ---")
            
            t0 = time.perf_counter()
            sol, init_cost, final_cost, is_valid = solve_instance(
                inst,
                budget_s=budget,
                validate=validate,
                verbose=verbose,
            )
            wall_time = time.perf_counter() - t0
            
            # Calculate metrics
            improvement_pct = ((init_cost - final_cost) / init_cost * 100) if init_cost > 0 else 0
            
            result = {
                'instance': inst_name,
                'n_nodes': inst.n_nodes,
                'n_vehicles': inst.n_vehicles,
                'capacity': inst.capacity,
                'budget': budget,
                'init_cost': init_cost,
                'final_cost': final_cost,
                'improvement_pct': improvement_pct,
                'n_routes': sol.num_routes(),
                'wall_time': wall_time,
                'is_valid': is_valid,
            }
            
            results.append(result)
            
            print(f"  Initial cost: {init_cost}")
            print(f"  Final cost: {final_cost}")
            print(f"  Improvement: {improvement_pct:.2f}%")
            print(f"  Routes: {sol.num_routes()}")
            print(f"  Wall time: {wall_time:.2f}s")
            print(f"  Valid: {'✓' if is_valid else '✗'}")
    
    # Write to Excel
    print(f"\n{'='*80}")
    print(f"Writing results to {output_excel}...")
    _write_excel(results, output_excel, time_budgets)
    print(f"✅ Complete!")


def _write_excel(results: List[Dict], output_path: Path, time_budgets: List[int]) -> None:
    """Write results to Excel with formatting."""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets for each time budget
    for budget in time_budgets:
        budget_results = [r for r in results if r['budget'] == budget]
        
        ws = wb.create_sheet(f"{budget}s")
        
        # Headers
        headers = ['Instance', 'Nodes', 'Vehicles', 'Capacity', 'Initial Cost', 
                   'Final Cost', 'Improvement %', 'Routes', 'Wall Time (s)', 'Valid']
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, result in enumerate(budget_results, 2):
            ws.cell(row=row_idx, column=1, value=result['instance'])
            ws.cell(row=row_idx, column=2, value=result['n_nodes'])
            ws.cell(row=row_idx, column=3, value=result['n_vehicles'])
            ws.cell(row=row_idx, column=4, value=result['capacity'])
            ws.cell(row=row_idx, column=5, value=result['init_cost'])
            ws.cell(row=row_idx, column=6, value=result['final_cost'])
            ws.cell(row=row_idx, column=7, value=f"{result['improvement_pct']:.2f}")
            ws.cell(row=row_idx, column=8, value=result['n_routes'])
            ws.cell(row=row_idx, column=9, value=f"{result['wall_time']:.2f}")
            ws.cell(row=row_idx, column=10, value='✓' if result['is_valid'] else '✗')
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = "Benchmark Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)
    
    ws_summary['A3'] = "Time Budgets:"
    ws_summary['B3'] = ", ".join(f"{b}s" for b in time_budgets)
    
    ws_summary['A4'] = "Instances:"
    ws_summary['B4'] = len(set(r['instance'] for r in results))
    
    ws_summary['A5'] = "Total Runs:"
    ws_summary['B5'] = len(results)
    
    ws_summary['A6'] = "All Valid:"
    ws_summary['B6'] = '✓' if all(r['is_valid'] for r in results) else '✗'
    
    # Budget comparison
    ws_summary['A8'] = "Average Improvement by Budget:"
    ws_summary['A8'].font = Font(bold=True)
    
    for idx, budget in enumerate(time_budgets, 9):
        budget_results = [r for r in results if r['budget'] == budget]
        avg_improvement = sum(r['improvement_pct'] for r in budget_results) / len(budget_results)
        ws_summary[f'A{idx}'] = f"{budget}s:"
        ws_summary[f'B{idx}'] = f"{avg_improvement:.2f}%"
    
    wb.save(output_path)
