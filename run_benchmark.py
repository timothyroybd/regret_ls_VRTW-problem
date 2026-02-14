#!/usr/bin/env python3
"""
Main script to run benchmark on all instances with detailed validation in Excel.

Usage:
    python run_benchmark.py
    python run_benchmark.py --budgets 300 600
    python run_benchmark.py --verbose
"""
import argparse
import sys
from pathlib import Path

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / 'src'))

from instance import load_ortec_vrptw
from solver import solve_instance
from validation import validate_solution
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Dict


def run_benchmark(
    instance_paths: List[Path],
    time_budgets: List[int],
    output_excel: Path,
    validate: bool = True,
    verbose: bool = False,
) -> None:
    """Run benchmark on multiple instances with multiple time budgets."""
    results = []
    
    for inst_path in instance_paths:
        inst_name = inst_path.stem
        print(f"\n{'='*80}")
        print(f"Instance: {inst_name}")
        print(f"{'='*80}")
        
        # Load instance
        print(f"Loading instance...")
        inst = load_ortec_vrptw(inst_path)
        print(f"  Nodes: {inst.n_nodes}, Vehicles: {inst.n_vehicles}, Capacity: {inst.capacity}")
        
        for budget in time_budgets:
            print(f"\n--- Budget: {budget}s ---")
            
            t0 = time.perf_counter()
            sol, init_cost, final_cost, is_valid = solve_instance(
                inst,
                budget_s=budget,
                validate=validate,
                verbose=verbose,
            )
            wall_time = time.perf_counter() - t0
            
            # Get detailed validation info
            validation_result = validate_solution(inst, sol, verbose=False)
            
            # Calculate metrics
            improvement_pct = ((init_cost - final_cost) / init_cost * 100) if init_cost > 0 else 0
            
            # Calculate average capacity utilization
            total_capacity_used = sum(rv.total_demand for rv in validation_result.route_validations)
            total_capacity_available = len(sol.routes) * inst.capacity
            capacity_utilization = (total_capacity_used / total_capacity_available * 100) if total_capacity_available > 0 else 0
            
            result = {
                'instance': inst_name,
                'n_nodes': inst.n_nodes,
                'n_vehicles': inst.n_vehicles,
                'capacity': inst.capacity,
                'budget': budget,
                'init_cost': init_cost,
                'final_cost': final_cost,
                'improvement_pct': improvement_pct,
                'n_routes': sol.num_routes(),
                'wall_time': wall_time,
                'is_valid': is_valid,
                'total_violations': validation_result.total_violations,
                'unrouted_customers': len(validation_result.unrouted_customers),
                'duplicate_customers': len(validation_result.duplicate_customers),
                'capacity_utilization': capacity_utilization,
                'customers_served': sol.num_customers(),
            }
            
            results.append(result)
            
            print(f"  Initial cost: {init_cost}")
            print(f"  Final cost: {final_cost}")
            print(f"  Improvement: {improvement_pct:.2f}%")
            print(f"  Routes: {sol.num_routes()}")
            print(f"  Customers served: {sol.num_customers()}/{inst.n_nodes - 1}")
            print(f"  Capacity utilization: {capacity_utilization:.1f}%")
            print(f"  Wall time: {wall_time:.2f}s")
            print(f"  Valid: {'✓' if is_valid else '✗'}")
            if validation_result.total_violations > 0:
                print(f"  ⚠️ Violations: {validation_result.total_violations}")
    
    # Write to Excel
    print(f"\n{'='*80}")
    print(f"Writing results to {output_excel}...")
    _write_excel(results, output_excel, time_budgets)
    print(f"✅ Complete!")


def _write_excel(results: List[Dict], output_path: Path, time_budgets: List[int]) -> None:
    """Write results to Excel with formatting and validation details."""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets for each time budget
    for budget in time_budgets:
        budget_results = [r for r in results if r['budget'] == budget]
        
        ws = wb.create_sheet(f"{budget}s Results")
        
        # Headers
        headers = [
            'Instance', 'Nodes', 'Vehicles', 'Capacity', 
            'Initial Cost', 'Final Cost', 'Improvement %', 
            'Routes Used', 'Customers Served', 'Capacity Util %',
            'Wall Time (s)', 'Valid', 'Violations', 'Unrouted', 'Duplicates'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, result in enumerate(budget_results, 2):
            ws.cell(row=row_idx, column=1, value=result['instance'])
            ws.cell(row=row_idx, column=2, value=result['n_nodes'])
            ws.cell(row=row_idx, column=3, value=result['n_vehicles'])
            ws.cell(row=row_idx, column=4, value=result['capacity'])
            ws.cell(row=row_idx, column=5, value=result['init_cost'])
            ws.cell(row=row_idx, column=6, value=result['final_cost'])
            ws.cell(row=row_idx, column=7, value=f"{result['improvement_pct']:.2f}")
            ws.cell(row=row_idx, column=8, value=result['n_routes'])
            ws.cell(row=row_idx, column=9, value=result['customers_served'])
            ws.cell(row=row_idx, column=10, value=f"{result['capacity_utilization']:.1f}")
            ws.cell(row=row_idx, column=11, value=f"{result['wall_time']:.2f}")
            
            # Validation columns with conditional formatting
            valid_cell = ws.cell(row=row_idx, column=12, value='✓' if result['is_valid'] else '✗')
            if not result['is_valid']:
                valid_cell.font = Font(color="FF0000", bold=True)
            
            violations_cell = ws.cell(row=row_idx, column=13, value=result['total_violations'])
            if result['total_violations'] > 0:
                violations_cell.font = Font(color="FF0000", bold=True)
            
            unrouted_cell = ws.cell(row=row_idx, column=14, value=result['unrouted_customers'])
            if result['unrouted_customers'] > 0:
                unrouted_cell.font = Font(color="FF0000", bold=True)
            
            duplicates_cell = ws.cell(row=row_idx, column=15, value=result['duplicate_customers'])
            if result['duplicate_customers'] > 0:
                duplicates_cell.font = Font(color="FF0000", bold=True)
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Validation Summary Sheet
    ws_val = wb.create_sheet("Validation Summary", 0)
    ws_val['A1'] = "VALIDATION SUMMARY"
    ws_val['A1'].font = Font(bold=True, size=14)
    
    ws_val['A3'] = "Overall Validation Status"
    ws_val['A3'].font = Font(bold=True)
    
    all_valid = all(r['is_valid'] for r in results)
    ws_val['A4'] = "All Solutions Valid:"
    ws_val['B4'] = '✓ YES' if all_valid else '✗ NO'
    if not all_valid:
        ws_val['B4'].font = Font(color="FF0000", bold=True)
    else:
        ws_val['B4'].font = Font(color="00B050", bold=True)
    
    ws_val['A6'] = "Validation Metrics by Budget"
    ws_val['A6'].font = Font(bold=True)
    
    row = 7
    for budget in time_budgets:
        budget_results = [r for r in results if r['budget'] == budget]
        
        valid_count = sum(1 for r in budget_results if r['is_valid'])
        total_violations = sum(r['total_violations'] for r in budget_results)
        total_unrouted = sum(r['unrouted_customers'] for r in budget_results)
        total_duplicates = sum(r['duplicate_customers'] for r in budget_results)
        avg_capacity_util = sum(r['capacity_utilization'] for r in budget_results) / len(budget_results)
        
        ws_val[f'A{row}'] = f"{budget}s Budget:"
        ws_val[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws_val[f'A{row}'] = "  Valid Solutions:"
        ws_val[f'B{row}'] = f"{valid_count}/{len(budget_results)}"
        row += 1
        
        ws_val[f'A{row}'] = "  Total Violations:"
        ws_val[f'B{row}'] = total_violations
        if total_violations > 0:
            ws_val[f'B{row}'].font = Font(color="FF0000")
        row += 1
        
        ws_val[f'A{row}'] = "  Unrouted Customers:"
        ws_val[f'B{row}'] = total_unrouted
        if total_unrouted > 0:
            ws_val[f'B{row}'].font = Font(color="FF0000")
        row += 1
        
        ws_val[f'A{row}'] = "  Duplicate Customers:"
        ws_val[f'B{row}'] = total_duplicates
        if total_duplicates > 0:
            ws_val[f'B{row}'].font = Font(color="FF0000")
        row += 1
        
        ws_val[f'A{row}'] = "  Avg Capacity Utilization:"
        ws_val[f'B{row}'] = f"{avg_capacity_util:.1f}%"
        row += 2
    
    # Performance Summary Sheet
    ws_summary = wb.create_sheet("Performance Summary", 1)
    ws_summary['A1'] = "BENCHMARK SUMMARY"
    ws_summary['A1'].font = Font(bold=True, size=14)
    
    ws_summary['A3'] = "Time Budgets:"
    ws_summary['B3'] = ", ".join(f"{b}s" for b in time_budgets)
    
    ws_summary['A4'] = "Instances:"
    ws_summary['B4'] = len(set(r['instance'] for r in results))
    
    ws_summary['A5'] = "Total Runs:"
    ws_summary['B5'] = len(results)
    
    ws_summary['A7'] = "Average Improvement by Budget:"
    ws_summary['A7'].font = Font(bold=True)
    
    for idx, budget in enumerate(time_budgets, 8):
        budget_results = [r for r in results if r['budget'] == budget]
        avg_improvement = sum(r['improvement_pct'] for r in budget_results) / len(budget_results)
        ws_summary[f'A{idx}'] = f"{budget}s:"
        ws_summary[f'B{idx}'] = f"{avg_improvement:.2f}%"
    
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description='Run VRPTW benchmark')
    parser.add_argument(
        '--budgets',
        type=int,
        nargs='+',
        default=[300, 600],
        help='Time budgets in seconds (default: 300 600)'
    )
    parser.add_argument(
        '--instances',
        type=str,
        default='instances',
        help='Path to instances directory (default: instances)'
    )
    parser.add_argument(
        '--output',
        type=str,
        default='results/benchmark_results.xlsx',
        help='Output Excel file (default: results/benchmark_results.xlsx)'
    )
    parser.add_argument(
        '--no-validate',
        action='store_true',
        help='Skip validation (faster but not recommended)'
    )
    parser.add_argument(
        '--verbose',
        action='store_true',
        help='Verbose output'
    )
    
    args = parser.parse_args()
    
    # Get instance files
    instance_dir = Path(args.instances)
    if not instance_dir.exists():
        print(f"Error: Instance directory not found: {instance_dir}")
        return 1
    
    instance_paths = sorted(instance_dir.glob('ORTEC-VRPTW-*.txt'))
    
    if not instance_paths:
        print(f"Error: No instance files found in {instance_dir}")
        return 1
    
    print(f"Found {len(instance_paths)} instances")
    print(f"Time budgets: {args.budgets}")
    print(f"Output: {args.output}")
    print(f"Validation: {'OFF' if args.no_validate else 'ON'}")
    
    # Create output directory
    output_path = Path(args.output)
    output_path.parent.mkdir(exist_ok=True, parents=True)
    
    # Run benchmark
    run_benchmark(
        instance_paths=instance_paths,
        time_budgets=args.budgets,
        output_excel=output_path,
        validate=not args.no_validate,
        verbose=args.verbose,
    )
    
    return 0


if __name__ == '__main__':
    sys.exit(main())
